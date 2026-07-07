import csv
from pathlib import Path

from django.utils import timezone
from openpyxl import load_workbook

from ..models import Job

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
PREVIEW_ROW_LIMIT = 25


def mark_job_running(job: Job) -> None:
    job.status = Job.Status.RUNNING
    job.error_message = ""
    job.num_processed = 0
    job.changed_row_count = 0
    job.started_at = timezone.now()
    job.save(
        update_fields=[
            "status",
            "error_message",
            "num_processed",
            "changed_row_count",
            "started_at",
        ]
    )


def mark_job_success(job: Job) -> None:
    job.status = Job.Status.SUCCESS
    job.error_message = ""
    job.completed_at = timezone.now()
    job.save(update_fields=["status", "error_message", "completed_at"])


def mark_job_failed(job: Job, error_message: str) -> None:
    job.status = Job.Status.FAILED
    job.error_message = error_message
    job.completed_at = timezone.now()
    job.save(update_fields=["status", "error_message", "completed_at"])


def validate_input_file(job: Job) -> str:
    input_path = job.input_file.path
    extension = Path(input_path).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            "Unsupported file type. Only CSV and XLSX files are supported."
        )

    return input_path


def collect_csv_metadata(file_path: str) -> tuple[int, list[str], list[dict[str, str]]]:
    with open(file_path, newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        headers = reader.fieldnames or []
        preview_rows: list[dict[str, str]] = []
        row_count = 0

        for row in reader:
            row_count += 1

            if len(preview_rows) < PREVIEW_ROW_LIMIT:
                preview_rows.append(
                    {
                        header: (value if value is not None else "")
                        for header, value in row.items()
                    }
                )

    return row_count, headers, preview_rows


def collect_xlsx_metadata(
    file_path: str,
) -> tuple[int, list[str], list[dict[str, str | int | float | bool]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if header_row is None:
            return 0, [], []

        headers = [
            str(value).strip() if value is not None else ""
            for value in header_row
        ]
        preview_rows: list[dict[str, str | int | float | bool]] = []
        row_count = 0

        for row in rows:
            row_count += 1

            if len(preview_rows) < PREVIEW_ROW_LIMIT:
                preview_rows.append(
                    {
                        headers[index]: (value if value is not None else "")
                        for index, value in enumerate(row)
                        if index < len(headers)
                    }
                )

        return row_count, headers, preview_rows
    finally:
        workbook.close()


def collect_file_metadata(
    file_path: str,
) -> tuple[int, list[str], list[dict[str, str | int | float | bool]]]:
    extension = Path(file_path).suffix.lower()

    if extension == ".csv":
        return collect_csv_metadata(file_path)

    if extension == ".xlsx":
        return collect_xlsx_metadata(file_path)

    raise ValueError("Unsupported file type. Only CSV and XLSX files are supported.")


def save_job_file_summary(
    job: Job,
    row_count: int,
    column_headers: list[str],
    preview_rows: list[dict[str, str]],
) -> None:
    job.row_count = row_count
    job.num_processed = row_count
    job.column_headers = column_headers
    job.preview_rows = preview_rows
    job.save(
        update_fields=[
            "row_count",
            "num_processed",
            "column_headers",
            "preview_rows",
        ]
    )
